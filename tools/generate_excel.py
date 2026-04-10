from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path
import tempfile

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference, ScatterChart, Series
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet import _writer as openpyxl_writer

ROOT_DIR = Path(__file__).resolve().parents[1]
OUTPUT_PATH = ROOT_DIR / "output" / "Ekologia_kalkulator.xlsx"
TEMP_DIR = ROOT_DIR / "tmp"
TEMP_DIR.mkdir(exist_ok=True)
tempfile.tempdir = str(TEMP_DIR)

_ORIGINAL_REMOVE = openpyxl_writer.os.remove


def _safe_remove(path):
    try:
        _ORIGINAL_REMOVE(path)
    except PermissionError:
        pass


openpyxl_writer.os.remove = _safe_remove

AIR_LIMITS = [
    ("benzen", 5),
    ("ditlenek azotu", 200),
    ("ditlenek siarki", 125),
    ("ozon", 120),
    ("pył zawieszony PM10", 50),
    ("tlenek węgla", 10000),
]

AIR_TOXICITY = [
    ("tlenek siarki IV", 1.0),
    ("tlenek węgla II", 0.5),
    ("tlenki azotu", 2.0),
    ("ozon", 10.3),
    ("pył zawieszony", 2.9),
    ("chrom", 160.0),
    ("cynk", 16.9),
    ("kadm", 640.0),
    ("miedź", 106.7),
    ("ołów", 320.0),
    ("żelazo", 2.0),
]

AIR_SYNONYMS = [
    ("O3", "ozon", "ozon", "ozon", "µg/m3"),
    ("CO", "tlenek węgla", "tlenek węgla", "tlenek węgla II", "µg/m3"),
    ("NO2", "ditlenek azotu", "ditlenek azotu", "tlenki azotu", "µg/m3"),
    ("SO2", "ditlenek siarki", "ditlenek siarki", "tlenek siarki IV", "µg/m3"),
    ("PM10", "pył zawieszony PM10", "pył zawieszony PM10", "pył zawieszony", "µg/m3"),
    ("BENZEN", "benzen", "benzen", "", "µg/m3"),
    ("NOx", "tlenki azotu", "", "tlenki azotu", "µg/m3"),
    ("CHROM", "chrom", "", "chrom", "µg/m3"),
    ("CYNK", "cynk", "", "cynk", "µg/m3"),
    ("KADM", "kadm", "", "kadm", "µg/m3"),
    ("MIEDŹ", "miedź", "", "miedź", "µg/m3"),
    ("OŁÓW", "ołów", "", "ołów", "µg/m3"),
    ("ŻELAZO", "żelazo", "", "żelazo", "µg/m3"),
]

AIR_SAMPLE = {
    "start_date": date(2026, 3, 1),
    "hour": "08:00",
    "stations": [
        ("miejska", "komunikacyjna", "dolnośląskie", "ul. Wyb. J. Conrada-Korzeniowskiego 18, Wrocław", "Stacja Wrocław - śródmieście"),
        ("podmiejska", "tło", "dolnośląskie", "ul. Parkowa 3, Długołęka", "Stacja Długołęka"),
        ("pozamiejska", "tło", "dolnośląskie", "Jeleniów 12, teren referencyjny", "Stacja Jeleniów"),
    ],
    "pollutants": [
        ("O3", "ozon", "µg/m3", [65, 72, 80, 78, 74, 70, 68, 66, 64, 61], [58, 60, 66, 69, 67, 62, 60, 59, 56, 54], [50, 53, 57, 61, 59, 56, 54, 52, 49, 47]),
        ("SO2", "ditlenek siarki", "µg/m3", [21, 22, 23, 24, 24, 25, 23, 22, 21, 20], [17, 18, 19, 20, 20, 19, 18, 18, 17, 16], [13, 14, 15, 15, 16, 15, 14, 14, 13, 13]),
        ("NO2", "ditlenek azotu", "µg/m3", [88, 92, 96, 101, 98, 95, 90, 87, 84, 82], [61, 64, 67, 69, 70, 68, 65, 63, 61, 59], [39, 41, 43, 45, 44, 42, 41, 39, 38, 37]),
        ("PM10", "pył zawieszony PM10", "µg/m3", [29, 31, 34, 36, 35, 33, 30, 28, 27, 26], [23, 24, 26, 27, 28, 27, 25, 24, 23, 22], [18, 19, 21, 22, 21, 20, 19, 18, 17, 17]),
        ("CO", "tlenek węgla", "µg/m3", [1600, 1700, 1800, 1750, 1690, 1650, 1620, 1580, 1540, 1500], [1300, 1360, 1400, 1380, 1340, 1320, 1290, 1260, 1230, 1210], [900, 940, 980, 990, 960, 940, 920, 900, 880, 860]),
        ("", "", "µg/m3", ["", "", "", "", "", "", "", "", "", ""], ["", "", "", "", "", "", "", "", "", ""], ["", "", "", "", "", "", "", "", "", ""]),
    ],
}

SOIL_SAMPLE = {
    "active": (6.7, 6.8, 5.6, 5.5),
    "exchange": (6.9, 7.0, 4.8, 4.9),
    "hydrolytic": (4.2, 0.1, 1.75),
    "base": (25.0, 0.1, 7.5, 0.1),
    "hectare": (10000, 0.20, 1500),
}

THIN_BORDER = Border(
    left=Side(style="thin", color="D7D7C9"),
    right=Side(style="thin", color="D7D7C9"),
    top=Side(style="thin", color="D7D7C9"),
    bottom=Side(style="thin", color="D7D7C9"),
)

FILL_INPUT = PatternFill("solid", fgColor="DDEBF7")
FILL_STATIC = PatternFill("solid", fgColor="E7E6E6")
FILL_FORMULA = PatternFill("solid", fgColor="F2F2F2")
FILL_SECTION = PatternFill("solid", fgColor="D9EAD3")
FILL_WARN = PatternFill("solid", fgColor="FFF2CC")
FILL_TITLE = PatternFill("solid", fgColor="EAD7C0")

FONT_HEADER = Font(name="Segoe UI", size=11, bold=True, color="214E36")
FONT_TITLE = Font(name="Segoe UI", size=14, bold=True, color="24312A")
FONT_NORMAL = Font(name="Segoe UI", size=10)

ALIGN_TOP = Alignment(vertical="top", wrap_text=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_cell(cell, *, fill=None, font=None, alignment=None, border=THIN_BORDER, number_format=None):
    if fill is not None:
        cell.fill = fill
    if font is not None:
        cell.font = font
    if alignment is not None:
        cell.alignment = alignment
    if border is not None:
        cell.border = border
    if number_format is not None:
        cell.number_format = number_format


def section_title(ws, cell_ref: str, text: str):
    cell = ws[cell_ref]
    cell.value = text
    style_cell(cell, fill=FILL_SECTION, font=FONT_HEADER, alignment=ALIGN_CENTER)


def merge_title(ws, start: str, end: str, text: str):
    ws.merge_cells(f"{start}:{end}")
    cell = ws[start]
    cell.value = text
    style_cell(cell, fill=FILL_TITLE, font=FONT_TITLE, alignment=ALIGN_CENTER)


def set_headers(ws, row: int, labels: list[str]):
    for col, label in enumerate(labels, start=1):
        cell = ws.cell(row=row, column=col, value=label)
        style_cell(cell, fill=FILL_SECTION, font=FONT_HEADER, alignment=ALIGN_CENTER)


def set_widths(ws, widths: dict[str, float]):
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def data_range_formula(column: str, first_row: int, last_row: int, fn: str) -> str:
    return f'=IF(COUNTA({column}{first_row}:{column}{last_row})=0,"",{fn}({column}{first_row}:{column}{last_row}))'


def station_value_column(station_name: str) -> str:
    return {"miejska": "B", "podmiejska": "C", "pozamiejska": "D"}[station_name]


def air_block_start(slot_index: int) -> int:
    return 12 + slot_index * 15


def build_readme_sheet(ws):
    merge_title(ws, "A1", "F1", "Ekologia - instrukcja do workbooka")
    ws.freeze_panes = "A3"
    lines = [
        ("A3", "Ten workbook wspiera dwa ćwiczenia: poziom zanieczyszczenia powietrza oraz parametry fizykochemiczne gleby."),
        ("A4", "Kolory: niebieski = dane wejściowe, szary = stałe, jasnoszary = formuły, żółty = uwagi i ostrzeżenia."),
        ("A5", "Powietrze: wpisz alias zanieczyszczenia, nazwę, jednostkę i 10 kolejnych dni danych dla 3 stacji."),
        ("A6", "Gleba: wpisz surowe pomiary pH, objętości i stężenia. Arkusz wynikowy obliczy Hh, S, T, Mz, Hh_ha, CaO i CaCO3."),
        ("A8", "Wzory zaimplementowane w workbooku:"),
        ("A9", "Psi = Σ(Cx / D24x)"),
        ("A10", "Tau = Σ(Cx * alpha_t)"),
        ("A11", "Hh = V * c * 10 * k"),
        ("A12", "Mz = p * h * rho"),
        ("A13", "Hh_ha = Hh * Mz / 100"),
        ("A14", "CaO = Hh_ha * 0,028"),
        ("A15", "CaCO3 = Hh_ha * 0,050"),
        ("A16", "S = (V * c - V1 * c1) * 4 * 5"),
        ("A17", "T = Hh + S"),
        ("A19", "Założenie: klasyfikacja gleby korzysta ze średniego pH badanej próbki w KCl."),
        ("A20", "Brakujących limitów i współczynników toksyczności nie uzupełniono automatycznie - workbook pokazuje wtedy ostrzeżenia."),
    ]
    for cell_ref, value in lines:
        cell = ws[cell_ref]
        cell.value = value
        style_cell(cell, font=FONT_NORMAL, alignment=ALIGN_TOP)
    set_widths(ws, {"A": 96, "B": 20, "C": 20, "D": 20, "E": 20, "F": 20})


def build_air_data_sheet(ws):
    merge_title(ws, "A1", "H1", "Powietrze - dane wejściowe")
    ws.freeze_panes = "A6"
    ws["A2"] = "Data początkowa"
    ws["B2"] = AIR_SAMPLE["start_date"]
    ws["A3"] = "Data końcowa"
    ws["B3"] = "=B2+9"
    ws["A4"] = "Godzina pomiaru"
    ws["B4"] = AIR_SAMPLE["hour"]
    for cell_ref in ("A2", "A3", "A4"):
        style_cell(ws[cell_ref], fill=FILL_STATIC, font=FONT_HEADER)
    style_cell(ws["B2"], fill=FILL_INPUT, font=FONT_NORMAL, number_format="dd.mm.yyyy")
    style_cell(ws["B3"], fill=FILL_FORMULA, font=FONT_NORMAL, number_format="dd.mm.yyyy")
    style_cell(ws["B4"], fill=FILL_INPUT, font=FONT_NORMAL)

    set_headers(ws, 6, ["Typ obszaru", "Typ punktu", "Województwo", "Adres", "Nazwa stacji"])
    for row_index, station in enumerate(AIR_SAMPLE["stations"], start=7):
        values = [station[0], station[1], station[2], station[3], station[4]]
        for column_index, value in enumerate(values, start=1):
            cell = ws.cell(row=row_index, column=column_index, value=value)
            style_cell(cell, fill=FILL_INPUT, font=FONT_NORMAL, alignment=ALIGN_TOP)

    for slot_index, pollutant in enumerate(AIR_SAMPLE["pollutants"]):
        start_row = air_block_start(slot_index)
        ws.cell(start_row, 1, f"Zanieczyszczenie {slot_index + 1}")
        ws.cell(start_row, 2, pollutant[0])
        ws.cell(start_row, 3, pollutant[1])
        ws.cell(start_row, 4, pollutant[2])
        ws.cell(start_row, 5, "Wprowadź 10 kolejnych dni danych")
        set_headers(ws, start_row + 1, ["Data", "Miejska", "Podmiejska", "Pozamiejska"])
        for column in range(1, 6):
            style_cell(ws.cell(start_row, column), fill=FILL_STATIC if column in (1, 5) else FILL_INPUT, font=FONT_HEADER if column == 1 else FONT_NORMAL, alignment=ALIGN_TOP)
        for offset in range(10):
            row = start_row + 2 + offset
            date_cell = ws.cell(row=row, column=1, value=f"=$B$2+{offset}")
            style_cell(date_cell, fill=FILL_FORMULA, font=FONT_NORMAL, number_format="dd.mm.yyyy")
            for col_index, series in enumerate(pollutant[3:6], start=2):
                value = series[offset]
                cell = ws.cell(row=row, column=col_index, value=value if value != "" else None)
                style_cell(cell, fill=FILL_INPUT, font=FONT_NORMAL)
        ws.cell(start_row + 12, 1, "Godzina pomiaru")
        ws.cell(start_row + 12, 2, "=$B$4")
        style_cell(ws.cell(start_row + 12, 1), fill=FILL_STATIC, font=FONT_HEADER)
        style_cell(ws.cell(start_row + 12, 2), fill=FILL_FORMULA, font=FONT_NORMAL)

    set_widths(ws, {"A": 20, "B": 18, "C": 22, "D": 22, "E": 36, "F": 18, "G": 18, "H": 18})


def build_air_mapping_sheet(ws):
    merge_title(ws, "A1", "K1", "Powietrze - mapowanie")
    ws.freeze_panes = "A3"
    set_headers(ws, 3, ["Substancja", "Limit D24 [µg/m3]"])
    for row, (name, value) in enumerate(AIR_LIMITS, start=4):
        ws.cell(row, 1, name)
        ws.cell(row, 2, value)
        style_cell(ws.cell(row, 1), fill=FILL_INPUT, font=FONT_NORMAL)
        style_cell(ws.cell(row, 2), fill=FILL_INPUT, font=FONT_NORMAL)

    set_headers(ws, 3, ["Zanieczyszczenie", "Współczynnik toksyczności"])
    ws["D3"] = "Zanieczyszczenie"
    ws["E3"] = "Współczynnik toksyczności"
    style_cell(ws["D3"], fill=FILL_SECTION, font=FONT_HEADER, alignment=ALIGN_CENTER)
    style_cell(ws["E3"], fill=FILL_SECTION, font=FONT_HEADER, alignment=ALIGN_CENTER)
    for row, (name, value) in enumerate(AIR_TOXICITY, start=4):
        ws.cell(row, 4, name)
        ws.cell(row, 5, value)
        style_cell(ws.cell(row, 4), fill=FILL_INPUT, font=FONT_NORMAL)
        style_cell(ws.cell(row, 5), fill=FILL_INPUT, font=FONT_NORMAL)

    for column_index, label in enumerate(["Alias", "Nazwa wyświetlana", "Nazwa do limitu", "Nazwa do toksyczności", "Jednostka"], start=7):
        cell = ws.cell(3, column_index, label)
        style_cell(cell, fill=FILL_SECTION, font=FONT_HEADER, alignment=ALIGN_CENTER)
    for row, values in enumerate(AIR_SYNONYMS, start=4):
        for column_index, value in enumerate(values, start=7):
            cell = ws.cell(row, column_index, value)
            style_cell(cell, fill=FILL_INPUT, font=FONT_NORMAL)

    set_widths(ws, {"A": 28, "B": 18, "D": 28, "E": 22, "G": 16, "H": 28, "I": 28, "J": 28, "K": 16})


def build_air_results_sheet(ws):
    merge_title(ws, "A1", "U1", "Powietrze - wyniki")
    ws.freeze_panes = "A10"
    set_headers(ws, 2, ["Stacja", "Psi", "Interpretacja Psi", "Tau", "Interpretacja Tau", "Przekroczenia średniej", "Przekroczenia maksimum", "Pominięcia z Psi", "Pominięcia z Tau"])
    for row_index, station in enumerate(["miejska", "podmiejska", "pozamiejska"], start=3):
        ws.cell(row_index, 1, station)
        ws.cell(row_index, 2, f'=SUMIF($B$10:$B$27,A{row_index},$O$10:$O$27)')
        ws.cell(row_index, 3, f'=IF(B{row_index}="","Brak danych",IF(B{row_index}<=1,"jakość powietrza zadowalająca","wyższa wartość oznacza gorszą jakość powietrza"))')
        ws.cell(row_index, 4, f'=SUMIF($B$10:$B$27,A{row_index},$P$10:$P$27)')
        ws.cell(row_index, 5, f'=IF(D{row_index}="","Brak danych","niższa wartość oznacza lepszą jakość powietrza")')
        ws.cell(row_index, 6, f'=COUNTIFS($B$10:$B$27,A{row_index},$M$10:$M$27,"TAK")')
        ws.cell(row_index, 7, f'=COUNTIFS($B$10:$B$27,A{row_index},$N$10:$N$27,"TAK")')
        ws.cell(row_index, 8, f'=COUNTIFS($B$10:$B$27,A{row_index},$Q$10:$Q$27,"<>")')
        ws.cell(row_index, 9, f'=COUNTIFS($B$10:$B$27,A{row_index},$R$10:$R$27,"<>")')
        for column in range(1, 10):
            style_cell(ws.cell(row_index, column), fill=FILL_FORMULA if column != 1 else FILL_STATIC, font=FONT_NORMAL, alignment=ALIGN_TOP)

    headers = ["Slot", "Stacja", "Alias", "Nazwa", "Jednostka", "Nazwa do limitu", "Limit D24", "Nazwa do toksyczności", "Wsp. toksyczności", "MIN", "ŚREDNIA", "MAX", "Średnia > limit", "MAX > limit", "Składnik Psi", "Składnik Tau", "Ostrzeżenie limit", "Ostrzeżenie toksyczność", "Psi stacji", "Tau stacji", "Uwagi"]
    set_headers(ws, 9, headers)
    detail_row = 10
    for slot_index in range(6):
        block_row = air_block_start(slot_index)
        for station_name in ["miejska", "podmiejska", "pozamiejska"]:
            value_col = station_value_column(station_name)
            first_data_row = block_row + 2
            last_data_row = block_row + 11
            ws.cell(detail_row, 1, slot_index + 1)
            ws.cell(detail_row, 2, station_name)
            ws.cell(detail_row, 3, f"=Powietrze_Dane!$B${block_row}")
            ws.cell(detail_row, 4, f"=Powietrze_Dane!$C${block_row}")
            ws.cell(detail_row, 5, f"=Powietrze_Dane!$D${block_row}")
            ws.cell(detail_row, 6, f'=IF(C{detail_row}&D{detail_row}="","",IFERROR(INDEX(Powietrze_Mapowanie!$I$4:$I$40,MATCH(C{detail_row},Powietrze_Mapowanie!$G$4:$G$40,0)),IFERROR(INDEX(Powietrze_Mapowanie!$I$4:$I$40,MATCH(D{detail_row},Powietrze_Mapowanie!$G$4:$G$40,0)),IF(D{detail_row}<>"",D{detail_row},C{detail_row}))))')
            ws.cell(detail_row, 7, f'=IF(F{detail_row}="","",IFERROR(INDEX(Powietrze_Mapowanie!$B$4:$B$20,MATCH(F{detail_row},Powietrze_Mapowanie!$A$4:$A$20,0)),""))')
            ws.cell(detail_row, 8, f'=IF(C{detail_row}&D{detail_row}="","",IFERROR(INDEX(Powietrze_Mapowanie!$J$4:$J$40,MATCH(C{detail_row},Powietrze_Mapowanie!$G$4:$G$40,0)),IFERROR(INDEX(Powietrze_Mapowanie!$J$4:$J$40,MATCH(D{detail_row},Powietrze_Mapowanie!$G$4:$G$40,0)),IF(D{detail_row}<>"",D{detail_row},C{detail_row}))))')
            ws.cell(detail_row, 9, f'=IF(H{detail_row}="","",IFERROR(INDEX(Powietrze_Mapowanie!$E$4:$E$30,MATCH(H{detail_row},Powietrze_Mapowanie!$D$4:$D$30,0)),""))')
            ws.cell(detail_row, 10, data_range_formula(value_col, first_data_row, last_data_row, "MIN"))
            ws.cell(detail_row, 11, data_range_formula(value_col, first_data_row, last_data_row, "AVERAGE"))
            ws.cell(detail_row, 12, data_range_formula(value_col, first_data_row, last_data_row, "MAX"))
            ws.cell(detail_row, 13, f'=IF(G{detail_row}="","Brak limitu",IF(K{detail_row}="","Brak danych",IF(K{detail_row}>G{detail_row},"TAK","NIE")))')
            ws.cell(detail_row, 14, f'=IF(G{detail_row}="","Brak limitu",IF(L{detail_row}="","Brak danych",IF(L{detail_row}>G{detail_row},"TAK","NIE")))')
            ws.cell(detail_row, 15, f'=IF(AND(ISNUMBER(K{detail_row}),ISNUMBER(G{detail_row})),K{detail_row}/G{detail_row},"")')
            ws.cell(detail_row, 16, f'=IF(AND(ISNUMBER(K{detail_row}),ISNUMBER(I{detail_row})),K{detail_row}*I{detail_row},"")')
            ws.cell(detail_row, 17, f'=IF(AND(K{detail_row}<>"",G{detail_row}=""),"Brak wartości dopuszczalnej w instrukcji — zanieczyszczenie pominięto w indeksie jakości.","")')
            ws.cell(detail_row, 18, f'=IF(AND(K{detail_row}<>"",I{detail_row}=""),"Brak współczynnika toksyczności w instrukcji — zanieczyszczenie pominięto w indeksie toksyczności.","")')
            ws.cell(detail_row, 19, f'=SUMIF($B$10:$B$27,B{detail_row},$O$10:$O$27)')
            ws.cell(detail_row, 20, f'=SUMIF($B$10:$B$27,B{detail_row},$P$10:$P$27)')
            ws.cell(detail_row, 21, "Wskaźnik liczony dla całej stacji")
            for column in range(1, 22):
                fill = FILL_STATIC if column in (1, 2) else FILL_FORMULA
                style_cell(ws.cell(detail_row, column), fill=fill, font=FONT_NORMAL, alignment=ALIGN_TOP)
            detail_row += 1

    set_widths(ws, {"A": 8, "B": 14, "C": 12, "D": 26, "E": 14, "F": 24, "G": 12, "H": 24, "I": 14, "J": 11, "K": 11, "L": 11, "M": 15, "N": 15, "O": 12, "P": 12, "Q": 24, "R": 24, "S": 12, "T": 12, "U": 26})


def build_soil_data_sheet(ws):
    merge_title(ws, "A1", "E1", "Gleba - dane wejściowe")
    ws.freeze_panes = "A4"
    section_title(ws, "A3", "Kwasowość czynna")
    set_headers(ws, 4, ["Przedmiot analizy", "Pomiar 1", "Pomiar 2", "Średnia"])
    active_rows = [("Woda destylowana", SOIL_SAMPLE["active"][0], SOIL_SAMPLE["active"][1]), ("Badana próbka gleby", SOIL_SAMPLE["active"][2], SOIL_SAMPLE["active"][3])]
    for row_index, values in enumerate(active_rows, start=5):
        ws.cell(row_index, 1, values[0])
        ws.cell(row_index, 2, values[1])
        ws.cell(row_index, 3, values[2])
        ws.cell(row_index, 4, f'=IF(COUNTA(B{row_index}:C{row_index})=0,"",AVERAGE(B{row_index}:C{row_index}))')
        for column in range(1, 5):
            fill = FILL_INPUT if column in (2, 3) else (FILL_FORMULA if column == 4 else FILL_STATIC)
            style_cell(ws.cell(row_index, column), fill=fill, font=FONT_NORMAL)

    section_title(ws, "A9", "Kwasowość wymienna")
    set_headers(ws, 10, ["Przedmiot analizy", "Pomiar 1", "Pomiar 2", "Średnia"])
    exchange_rows = [("KCl", SOIL_SAMPLE["exchange"][0], SOIL_SAMPLE["exchange"][1]), ("Badana próbka gleby", SOIL_SAMPLE["exchange"][2], SOIL_SAMPLE["exchange"][3])]
    for row_index, values in enumerate(exchange_rows, start=11):
        ws.cell(row_index, 1, values[0])
        ws.cell(row_index, 2, values[1])
        ws.cell(row_index, 3, values[2])
        ws.cell(row_index, 4, f'=IF(COUNTA(B{row_index}:C{row_index})=0,"",AVERAGE(B{row_index}:C{row_index}))')
        for column in range(1, 5):
            fill = FILL_INPUT if column in (2, 3) else (FILL_FORMULA if column == 4 else FILL_STATIC)
            style_cell(ws.cell(row_index, column), fill=fill, font=FONT_NORMAL)

    section_title(ws, "A15", "Kwasowość hydrolityczna")
    set_headers(ws, 16, ["Parametr", "Wartość", "Jednostka"])
    hydrolityc_rows = [("VNaOH", SOIL_SAMPLE["hydrolytic"][0], "cm3"), ("c_NaOH", SOIL_SAMPLE["hydrolytic"][1], "mol/dm3"), ("k", SOIL_SAMPLE["hydrolytic"][2], "współczynnik")]
    for row_index, values in enumerate(hydrolityc_rows, start=17):
        for column_index, value in enumerate(values, start=1):
            ws.cell(row_index, column_index, value)
            style_cell(ws.cell(row_index, column_index), fill=FILL_INPUT if column_index == 2 else FILL_STATIC, font=FONT_NORMAL)

    section_title(ws, "A22", "Suma kationów zasadowych")
    set_headers(ws, 23, ["Parametr", "Wartość", "Jednostka"])
    base_rows = [("V_HCl", SOIL_SAMPLE["base"][0], "cm3"), ("c_HCl", SOIL_SAMPLE["base"][1], "mol/dm3"), ("V1_NaOH", SOIL_SAMPLE["base"][2], "cm3"), ("c1_NaOH", SOIL_SAMPLE["base"][3], "mol/dm3")]
    for row_index, values in enumerate(base_rows, start=24):
        for column_index, value in enumerate(values, start=1):
            ws.cell(row_index, column_index, value)
            style_cell(ws.cell(row_index, column_index), fill=FILL_INPUT if column_index == 2 else FILL_STATIC, font=FONT_NORMAL)

    section_title(ws, "A30", "Parametry dla 1 ha")
    set_headers(ws, 31, ["Parametr", "Wartość", "Jednostka"])
    hectare_rows = [("p", SOIL_SAMPLE["hectare"][0], "m2"), ("h", SOIL_SAMPLE["hectare"][1], "m"), ("rho", SOIL_SAMPLE["hectare"][2], "kg/m3")]
    for row_index, values in enumerate(hectare_rows, start=32):
        for column_index, value in enumerate(values, start=1):
            ws.cell(row_index, column_index, value)
            style_cell(ws.cell(row_index, column_index), fill=FILL_INPUT if column_index == 2 else FILL_STATIC, font=FONT_NORMAL)

    validation = DataValidation(type="list", formula1='"1.5,1.75,2.0"', allow_blank=False)
    ws.add_data_validation(validation)
    validation.add(ws["B19"])
    set_widths(ws, {"A": 28, "B": 16, "C": 16, "D": 16, "E": 30})


def build_soil_results_sheet(ws):
    merge_title(ws, "A1", "F1", "Gleba - wyniki")
    ws.freeze_panes = "A4"
    set_headers(ws, 3, ["Wynik", "Wartość", "Jednostka", "Uwagi"])
    rows = [
        ("Średnie pH wody destylowanej", '=IF(COUNTA(Gleba_Dane!B5:C5)=0,"",AVERAGE(Gleba_Dane!B5:C5))', "", "wartość pomocnicza"),
        ("Średnie pH próbki w wodzie", '=IF(COUNTA(Gleba_Dane!B6:C6)=0,"",AVERAGE(Gleba_Dane!B6:C6))', "", "kwasowość czynna"),
        ("Średnie pH KCl", '=IF(COUNTA(Gleba_Dane!B11:C11)=0,"",AVERAGE(Gleba_Dane!B11:C11))', "", "wartość pomocnicza"),
        ("Średnie pH próbki w KCl", '=IF(COUNTA(Gleba_Dane!B12:C12)=0,"",AVERAGE(Gleba_Dane!B12:C12))', "", "podstawa klasyfikacji"),
        ("Hh", '=IF(OR(Gleba_Dane!B17="",Gleba_Dane!B18="",Gleba_Dane!B19=""),"",Gleba_Dane!B17*Gleba_Dane!B18*10*Gleba_Dane!B19)', "mmol(+)/100 g", "kwasowość hydrolityczna"),
        ("Mz", '=IF(OR(Gleba_Dane!B32="",Gleba_Dane!B33="",Gleba_Dane!B34=""),"",Gleba_Dane!B32*Gleba_Dane!B33*Gleba_Dane!B34)', "kg", "masa gleby dla 1 ha"),
        ("Hh_ha", '=IF(OR(B8="",B9=""),"",B8*B9/100)', "mol(+)", "całkowita kwasowość hydrolityczna"),
        ("CaO", '=IF(B10="","",B10*0.028)', "kg/ha", "zapotrzebowanie na wapno"),
        ("CaCO3", '=IF(B10="","",B10*0.05)', "kg/ha", "zapotrzebowanie na wapno"),
        ("S", '=IF(OR(Gleba_Dane!B24="",Gleba_Dane!B25="",Gleba_Dane!B26="",Gleba_Dane!B27=""),"",(Gleba_Dane!B24*Gleba_Dane!B25-Gleba_Dane!B26*Gleba_Dane!B27)*4*5)', "mmol(+)/100 g", "suma kationów zasadowych"),
        ("T", '=IF(OR(B8="",B13=""),"",B8+B13)', "mmol(+)/100 g", "pojemność sorpcyjna"),
        ("Klasa gleby uprawnej", '=IF(B7="","Brak danych",IF(B7<=4,"bardzo kwaśne",IF(B7<=4.5,"kwaśne",IF(B7<=5,"średnio kwaśne",IF(B7<=6,"słabo kwaśne",IF(B7<=6.5,"obojętne",IF(B7<=7,"słabo alkaliczne",IF(B7<=7.5,"średnio alkaliczne","alkaliczne"))))))))', "", "klasyfikacja"),
        ("Klasa gleby leśnej", '=IF(B7="","Brak danych",IF(B7<=3.5,"bardzo silnie kwaśne",IF(B7<=4.5,"silnie kwaśne",IF(B7<=5.5,"kwaśne",IF(B7<=6.5,"słabo kwaśne",IF(B7<=7.2,"obojętne",IF(B7<=8,"słabo alkaliczne","alkaliczne")))))))', "", "klasyfikacja"),
    ]
    for row_index, values in enumerate(rows, start=4):
        for column_index, value in enumerate(values, start=1):
            ws.cell(row_index, column_index, value)
            fill = FILL_STATIC if column_index in (1, 3, 4) else FILL_FORMULA
            style_cell(ws.cell(row_index, column_index), fill=fill, font=FONT_NORMAL, alignment=ALIGN_TOP)

    ws["E4"] = "Interpretacja"
    ws["E5"] = "określ kategorię gleby"
    ws["E6"] = "odnieś pH KCl do klasy gleby uprawnej i leśnej"
    ws["E7"] = "skomentuj Hh, S i T"
    for cell_ref in ("E4", "E5", "E6", "E7"):
        style_cell(ws[cell_ref], fill=FILL_WARN, font=FONT_HEADER if cell_ref == "E4" else FONT_NORMAL, alignment=ALIGN_TOP)

    set_widths(ws, {"A": 34, "B": 18, "C": 18, "D": 24, "E": 34, "F": 18})


def build_air_charts_sheet(ws):
    merge_title(ws, "A1", "L1", "Powietrze - wykresy")
    ws["A2"] = "Wykresy korzystają bezpośrednio z danych z arkusza Powietrze_Dane."
    style_cell(ws["A2"], fill=FILL_WARN, font=FONT_NORMAL)

    anchors = ["A4", "G4", "A20", "G20", "A36", "G36"]
    for slot_index, anchor in enumerate(anchors):
        start_row = air_block_start(slot_index)
        alias = AIR_SAMPLE["pollutants"][slot_index][0] or f"Slot {slot_index + 1}"
        chart = ScatterChart()
        chart.title = f"Trend stężeń - {alias}"
        chart.style = 13
        chart.scatterStyle = "smoothMarker"
        chart.height = 7
        chart.width = 10
        chart.x_axis.title = "Data"
        chart.y_axis.title = "Stężenie [µg/m3]"
        xvalues = Reference(Workbook().active, min_col=1, min_row=1, max_row=1)
        xvalues = Reference(ws.parent["Powietrze_Dane"], min_col=1, min_row=start_row + 2, max_row=start_row + 11)
        for station_name, color in [("miejska", "2F6A48"), ("podmiejska", "3F6B85"), ("pozamiejska", "A57845")]:
            column = {"miejska": 2, "podmiejska": 3, "pozamiejska": 4}[station_name]
            values = Reference(ws.parent["Powietrze_Dane"], min_col=column, min_row=start_row + 2, max_row=start_row + 11)
            series = Series(values, xvalues, title=station_name)
            series.marker.symbol = "circle"
            series.marker.size = 6
            series.graphicalProperties.line.solidFill = color
            series.graphicalProperties.line.width = 24000
            chart.series.append(series)
        ws.add_chart(chart, anchor)
    set_widths(ws, {"A": 18, "B": 18, "C": 18, "D": 18, "E": 18, "F": 18, "G": 18, "H": 18, "I": 18, "J": 18, "K": 18, "L": 18})


def build_soil_charts_sheet(ws):
    merge_title(ws, "A1", "J1", "Gleba - wizualizacje pomocnicze")
    ws["A2"] = "Wizualizacje pomocnicze"
    style_cell(ws["A2"], fill=FILL_WARN, font=FONT_HEADER)
    ws["A4"] = "Porównanie pH"
    ws["A5"] = "Średnie pH wody destylowanej"
    ws["A6"] = "Średnie pH próbki w wodzie"
    ws["A7"] = "Średnie pH KCl"
    ws["A8"] = "Średnie pH próbki w KCl"
    for row in range(5, 9):
        ws.cell(row, 2, f"=Gleba_Wyniki!B{row}")
    ws["A12"] = "Hh vs S vs T"
    ws["A13"] = "Hh"
    ws["A14"] = "S"
    ws["A15"] = "T"
    ws["B13"] = "=Gleba_Wyniki!B8"
    ws["B14"] = "=Gleba_Wyniki!B13"
    ws["B15"] = "=Gleba_Wyniki!B14"
    ws["A19"] = "CaO vs CaCO3"
    ws["A20"] = "CaO"
    ws["A21"] = "CaCO3"
    ws["B20"] = "=Gleba_Wyniki!B11"
    ws["B21"] = "=Gleba_Wyniki!B12"

    for row in list(range(4, 9)) + list(range(12, 16)) + list(range(19, 22)):
        style_cell(ws.cell(row, 1), fill=FILL_STATIC, font=FONT_NORMAL)
        style_cell(ws.cell(row, 2), fill=FILL_FORMULA, font=FONT_NORMAL)

    def add_bar_chart(title: str, min_row: int, max_row: int, anchor: str):
        chart = BarChart()
        chart.title = title
        chart.style = 10
        chart.height = 7
        chart.width = 10
        chart.y_axis.title = "Wartość"
        data = Reference(ws, min_col=2, min_row=min_row, max_row=max_row)
        categories = Reference(ws, min_col=1, min_row=min_row, max_row=max_row)
        chart.add_data(data, titles_from_data=False)
        chart.set_categories(categories)
        ws.add_chart(chart, anchor)

    add_bar_chart("Wizualizacje pomocnicze - porównanie pH", 5, 8, "D4")
    add_bar_chart("Wizualizacje pomocnicze - Hh, S i T", 13, 15, "D20")
    add_bar_chart("Wizualizacje pomocnicze - CaO i CaCO3", 20, 21, "D36")
    set_widths(ws, {"A": 30, "B": 18, "C": 18, "D": 18, "E": 18, "F": 18, "G": 18, "H": 18, "I": 18, "J": 18})


def autofit_sheet(ws):
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        max_length = 0
        for cell in column_cells:
            value = cell.value
            if value is None:
                continue
            max_length = max(max_length, len(str(value)))
        if max_length and ws.column_dimensions[letter].width is None:
            ws.column_dimensions[letter].width = min(max_length + 2, 44)


def build_workbook() -> Workbook:
    wb = Workbook()
    readme_ws = wb.active
    readme_ws.title = "README"
    air_data_ws = wb.create_sheet("Powietrze_Dane")
    air_mapping_ws = wb.create_sheet("Powietrze_Mapowanie")
    air_results_ws = wb.create_sheet("Powietrze_Wyniki")
    air_charts_ws = wb.create_sheet("Powietrze_Wykresy")
    soil_data_ws = wb.create_sheet("Gleba_Dane")
    soil_results_ws = wb.create_sheet("Gleba_Wyniki")
    soil_charts_ws = wb.create_sheet("Gleba_Wykresy")

    build_readme_sheet(readme_ws)
    build_air_data_sheet(air_data_ws)
    build_air_mapping_sheet(air_mapping_ws)
    build_air_results_sheet(air_results_ws)
    build_air_charts_sheet(air_charts_ws)
    build_soil_data_sheet(soil_data_ws)
    build_soil_results_sheet(soil_results_ws)
    build_soil_charts_sheet(soil_charts_ws)

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True

    for ws in wb.worksheets:
        autofit_sheet(ws)
    return wb


def verify_workbook(path: Path):
    wb = load_workbook(path)
    expected = [
        "README",
        "Powietrze_Dane",
        "Powietrze_Mapowanie",
        "Powietrze_Wyniki",
        "Powietrze_Wykresy",
        "Gleba_Dane",
        "Gleba_Wyniki",
        "Gleba_Wykresy",
    ]
    assert wb.sheetnames == expected, wb.sheetnames
    assert wb["Powietrze_Wyniki"]["B3"].data_type == "f"
    assert wb["Gleba_Wyniki"]["B8"].data_type == "f"
    assert path.exists()


def main():
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    workbook = build_workbook()
    workbook.save(OUTPUT_PATH)
    verify_workbook(OUTPUT_PATH)
    print(f"Workbook zapisany: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
